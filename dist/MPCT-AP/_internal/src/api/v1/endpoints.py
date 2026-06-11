"""
src/api/v1/endpoints.py
========================
API v1 Route Handlers
Team: algoRoute | Project: MPCT-AP

PURPOSE
-------
This module defines every HTTP route in the v1 API surface:

  POST /api/v1/extract              → submit a multi-year extraction job;
                                       streams live SSE progress events
  GET  /api/v1/download/{session_id}→ fetch the compiled .xlsx after "complete"
  GET  /api/v1/health               → supervisor health check

ARCHITECTURAL UPGRADE — SSE + FILE HANDOFF
────────────────────────────────────────────
Previous design: POST /extract → long-poll → binary .xlsx response
New design:
  1. POST /extract opens a text/event-stream.  JSON events flow back live
     as the Supervisor works through each (year, month) pair.
  2. When done, a "complete" event carries a download URL:
       {"type":"complete", "url":"/api/v1/download/<session_id>"}
  3. GET /download/<session_id> returns the .xlsx binary from an in-memory
     store (keyed by session_id, TTL-expired by the reaper).

Since binary data cannot be mixed into a text/event-stream, the file
handoff is separated onto its own endpoint.

SSE EVENT TYPES
────────────────
  {"type":"status", "year":2024, "month":4,  "status":"processing"}
  {"type":"status", "year":2024, "month":4,  "status":"success"}
  {"type":"status", "year":2024, "month":4,  "status":"no_data"}
  {"type":"status", "year":2024, "month":4,  "status":"error", "message":"..."}
  {"type":"complete", "url":"/api/v1/download/<session_id>"}

MULTI-YEAR PAYLOAD
────────────────────
Request body now accepts:
  {
    "ifsc_code": "SBIN0000377",
    "account_number": "10554356145",
    "years_config": [
      {"year": 2024, "months": [4, 5, 12]},
      {"year": 2025, "months": [1, 2, 3]}
    ]
  }
"""

import asyncio
import io
import json
import logging
import os
import re
import subprocess
import sys
import time
import uuid
from pathlib import Path
from typing import AsyncGenerator

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel, Field, field_validator

from src.api.deps import (
    enforce_payload_size,
    get_supervisor,
    resolve_idempotency_hmac,
)
from src.automation.supervisor import (
    AutomationSupervisor,
    ExtractionJobMultiYear,
    QueueSaturatedError,
    SupervisorUnhealthyError,
)
from src.core.policies.extraction import (
    MAX_FISCAL_YEAR,
    MAX_MONTHS_PER_JOB,
    MIN_FISCAL_YEAR,
)
from src.core.policies.system import MAX_QUEUE_WAIT_SECONDS
from src.core.security import CanonicalizeError, canonicalize_account_number, canonicalize_ifsc

# ---------------------------------------------------------------------------
# IN-MEMORY DOWNLOAD STORE
# ---------------------------------------------------------------------------
# Maps session_id → {"bytes": <xlsx bytes>, "filename": str, "created_at": float}
# The TTL reaper in the Supervisor cleans up checkpoint files; here we apply
# our own lightweight expiry on GET /download so memory doesn't grow unbounded.
# TTL is intentionally generous (30 minutes) so the client has time to retry.
_DOWNLOAD_STORE: dict[str, dict] = {}
_DOWNLOAD_TTL_SECONDS: int = 1800   # 30 minutes


def _store_download(session_id: str, xlsx_bytes: bytes, filename: str) -> None:
    """Persist the compiled .xlsx bytes in the in-memory store."""
    # Prune expired entries on every write to prevent unbounded growth.
    now = time.time()
    expired = [k for k, v in _DOWNLOAD_STORE.items()
               if now - v["created_at"] > _DOWNLOAD_TTL_SECONDS]
    for k in expired:
        del _DOWNLOAD_STORE[k]

    _DOWNLOAD_STORE[session_id] = {
        "bytes":      xlsx_bytes,
        "filename":   filename,
        "created_at": now,
    }


def get_executable_dir() -> str:
    """Get the directory where the actual .exe file lives, or the repository root when running raw."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.abspath(
        os.path.join(os.path.dirname(__file__), os.pardir, os.pardir, os.pardir)
    )


def _open_folder(path: str) -> None:
    """Open the file or folder in the OS file manager if possible."""
    try:
        if sys.platform == "win32":
            try:
                if os.path.isfile(path):
                    subprocess.Popen(
                        ["explorer", "/select,", os.path.abspath(path)],
                        shell=False,
                        creationflags=subprocess.CREATE_NO_WINDOW,
                    )
                else:
                    import ctypes
                    ctypes.windll.shell32.ShellExecuteW(None, "open", path, None, None, 1)
                return
            except Exception:
                pass

            # Fallback to explorer if ShellExecute fails.
            if os.path.isfile(path):
                subprocess.Popen(
                    ["explorer", "/select,", os.path.abspath(path)],
                    shell=False,
                    creationflags=subprocess.CREATE_NO_WINDOW,
                )
            else:
                subprocess.Popen(
                    ["explorer", path],
                    shell=False,
                    creationflags=subprocess.CREATE_NO_WINDOW,
                )
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path], shell=False)
        else:
            subprocess.Popen(["xdg-open", path], shell=False)
    except Exception as exc:
        logger.warning("[ENDPOINT] Could not open results folder %s: %s", path, exc)


def _get_desktop_folder() -> str:
    """Return the current user's desktop folder path on Windows or fallback."""
    if sys.platform == "win32":
        try:
            import ctypes.wintypes

            CSIDL_DESKTOPDIRECTORY = 0x0010
            buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
            ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOPDIRECTORY, None, 0, buf)
            desktop_folder = buf.value
            if desktop_folder:
                return desktop_folder
        except Exception:
            pass

    # Fallback to the standard Desktop path if the Windows API call is unavailable.
    home = Path.home()
    desktop_folder = os.path.join(home, "Desktop")
    if os.path.isdir(desktop_folder):
        return desktop_folder

    # Many Windows users have Desktop redirected into OneDrive.
    onedrive_desktop = os.path.join(home, "OneDrive", "Desktop")
    if os.path.isdir(onedrive_desktop):
        return onedrive_desktop

    try:
        os.makedirs(desktop_folder, exist_ok=True)
        return desktop_folder
    except Exception:
        pass

    # Final fallback: use the user's home directory if Desktop is missing/uncreatable.
    return str(home)


def _sanitize_filename_component(value: str, fallback: str) -> str:
    """Convert arbitrary text into a safe filename fragment."""
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", str(value).strip()).strip("_")
    cleaned = re.sub(r"_+", "_", cleaned)
    return cleaned[:60] or fallback


def _pick_primary_party_name(df: pd.DataFrame) -> str:
    """Pick the most representative party name from the exported rows."""
    for column_name in ("party_name", "Party Name"):
        if column_name not in df.columns:
            continue

        series = df[column_name].dropna().astype(str).str.strip()
        series = series[series != ""]
        if series.empty:
            continue

        return series.value_counts().idxmax()

    return "party"


def _build_unique_desktop_path(desktop_folder: str, filename: str) -> str:
    """Return a desktop path that does not overwrite an existing file."""
    candidate = os.path.join(desktop_folder, filename)
    if not os.path.exists(candidate):
        return candidate

    stem, suffix = os.path.splitext(filename)
    counter = 1
    while True:
        candidate = os.path.join(desktop_folder, f"{stem}_{counter}{suffix}")
        if not os.path.exists(candidate):
            return candidate
        counter += 1


def save_dataframe_to_local_results(df, filename="treasury_report.xlsx"):
    """Saves the Excel file directly to the user's Desktop folder with auto-sized columns."""

    desktop_folder = _get_desktop_folder()
    os.makedirs(desktop_folder, exist_ok=True)

    full_path = _build_unique_desktop_path(desktop_folder, filename)
    logger.info("[ENDPOINT] Saving Excel to Desktop path: %s", full_path)

    # Use ExcelWriter to format columns
    with pd.ExcelWriter(full_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Report")
        worksheet = writer.sheets["Report"]
        
        # Auto-adjust column widths
        for i, col in enumerate(df.columns):
            # Calculate max length of the data or the header title, whichever is longer
            max_len = max(
                df[col].astype(str).map(len).max() if not df[col].empty else 0, 
                len(str(col))
            )
            
            # openpyxl columns are 1-indexed (A=1, B=2...)
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(i + 1)
            
            # Set the width with extra padding for wide date characters
            worksheet.column_dimensions[col_letter].width = max_len + 8

    return full_path

logger = logging.getLogger(__name__)

# Create the v1 API router.  All routes defined here are mounted under
# /api/v1 by main.py using: app.include_router(router, prefix="/api/v1")
router = APIRouter(tags=["extraction"])


# ──────────────────────────────────────────────────────────────────────────────
# REQUEST / RESPONSE MODELS  (Pydantic)
# ──────────────────────────────────────────────────────────────────────────────

class YearConfig(BaseModel):
    """
    A single year + month-list pairing inside the multi-year payload.

    Example: {"year": 2024, "months": [4, 5, 6]}
    """
    year   : int = Field(
        ...,
        ge=MIN_FISCAL_YEAR,
        le=MAX_FISCAL_YEAR,
        description=f"Fiscal year ({MIN_FISCAL_YEAR}–{MAX_FISCAL_YEAR})",
    )
    months : list[int] = Field(
        default=list(range(1, 13)),
        description="1-based month numbers for this year",
    )

    @field_validator("months")
    @classmethod
    def validate_months(cls, v: list[int]) -> list[int]:
        if not v:
            raise ValueError("At least one month must be specified per year.")
        if len(v) > MAX_MONTHS_PER_JOB:
            raise ValueError(
                f"Too many months requested ({len(v)}); "
                f"maximum is {MAX_MONTHS_PER_JOB}."
            )
        for m in v:
            if not (1 <= m <= 12):
                raise ValueError(f"Month {m} is out of range (1–12).")
        return sorted(set(v))


class ExtractionRequest(BaseModel):
    """
    Multi-year extraction request body.

    Replaces the old single-year  year: int + months: list[int]  schema.
    The new `years_config` field carries an array of year-month clusters
    so a single request can span multiple financial years.

    Example payload:
      {
        "ifsc_code": "SBIN0000377",
        "account_number": "10554356145",
        "years_config": [
          {"year": 2024, "months": [4, 5, 12]},
          {"year": 2025, "months": [1, 2, 3]}
        ]
      }
    """
    ifsc_code      : str = Field(
        ...,
        description="Bank branch IFSC code (e.g. SBIN0001234)",
        examples=["SBIN0001234"],
    )
    account_number : str = Field(
        ...,
        description="Bank account number (9–18 digits)",
        examples=["123456789012"],
    )
    years_config   : list[YearConfig] = Field(
        ...,
        description="Array of year+month clusters to extract",
        min_length=1,
    )

    @field_validator("ifsc_code")
    @classmethod
    def validate_ifsc(cls, v: str) -> str:
        """Canonicalize and validate the IFSC code at model construction time."""
        try:
            return canonicalize_ifsc(v)
        except CanonicalizeError as exc:
            raise ValueError(str(exc)) from exc

    @field_validator("account_number")
    @classmethod
    def validate_account(cls, v: str) -> str:
        """Canonicalize the account number (strip non-digits, validate length)."""
        try:
            return canonicalize_account_number(v)
        except CanonicalizeError as exc:
            raise ValueError(str(exc)) from exc

    @field_validator("years_config")
    @classmethod
    def validate_years_config(cls, v: list[YearConfig]) -> list[YearConfig]:
        if not v:
            raise ValueError("At least one year configuration must be provided.")
        # Deduplicate years (keep last occurrence wins for months).
        seen: dict[int, YearConfig] = {}
        for yc in v:
            seen[yc.year] = yc
        return list(seen.values())


class HealthResponse(BaseModel):
    status              : str
    circuit_state       : str
    queue_depth         : int
    active_jobs         : int
    consecutive_crashes : int
    total_jobs_done     : int
    total_jobs_failed   : int
    uptime_seconds      : float


# ──────────────────────────────────────────────────────────────────────────────
# POST /extract  →  SSE live progress stream
# ──────────────────────────────────────────────────────────────────────────────

@router.post(
    "/extract",
    summary="Submit a multi-year extraction job; receive live SSE progress events",
    response_description="text/event-stream of JSON status events, ending with a download URL",
    responses={
        200: {"content": {"text/event-stream": {}}},
        413: {"description": "Request body too large"},
        422: {"description": "Validation error"},
        429: {"description": "Job queue is full"},
        503: {"description": "Supervisor circuit breaker open"},
    },
)
async def submit_extraction(
    request    : Request,
    body       : ExtractionRequest,
    _size      : None = Depends(enforce_payload_size),
    dedup_key  : str  = Depends(resolve_idempotency_hmac),
    supervisor : AutomationSupervisor = Depends(get_supervisor),
) -> StreamingResponse:
    """
    SSE-based extraction endpoint.

    Flow:
      1. Validate + canonicalize the multi-year body.
      2. Create an asyncio.Queue used by the Supervisor to push per-month
         status events back to this handler.
      3. Submit an ExtractionJobMultiYear to the Supervisor.
      4. Yield SSE events from the queue until the Supervisor pushes a
         sentinel (None) signalling job completion.
      5. Serialize the final DataFrame to .xlsx, store it in _DOWNLOAD_STORE
         keyed by session_id, then emit a "complete" event with the URL.

    SSE FORMAT (per the HTML spec):
        data: {"type": "status", ...}\n\n
    Each message is a single `data:` line followed by a blank line.
    """
    task_id    = f"req-{uuid.uuid4().hex[:10]}"
    session_id = uuid.uuid4().hex
    years_summary = [(yc.year, yc.months) for yc in body.years_config]

    logger.info(
        "[ENDPOINT] POST /extract (SSE)  task_id=%s  ifsc=%s  years=%s",
        task_id, body.ifsc_code, years_summary,
    )

    # ------------------------------------------------------------------
    # The progress_queue is the real-time pipe between the Supervisor's
    # worker loop and this streaming generator.
    #
    # Supervisor pushes:  dict event   → one SSE event is emitted
    #                     None         → sentinel; loop exits
    # ------------------------------------------------------------------
    progress_queue: asyncio.Queue = asyncio.Queue()

    loop   = asyncio.get_event_loop()
    future : asyncio.Future = loop.create_future()

    job = ExtractionJobMultiYear(
        task_id        = task_id,
        ifsc           = body.ifsc_code,
        account_no     = body.account_number,
        years_config   = [(yc.year, yc.months) for yc in body.years_config],
        result_future  = future,
        progress_queue = progress_queue,
    )

    try:
        await supervisor.submit(job)
    except QueueSaturatedError as exc:
        logger.warning("[ENDPOINT] Queue saturated for task %s: %s", task_id, exc)
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="The extraction queue is currently full. Please retry in a few seconds.",
            headers={"Retry-After": str(MAX_QUEUE_WAIT_SECONDS)},
        )
    except SupervisorUnhealthyError as exc:
        logger.error("[ENDPOINT] Supervisor unhealthy for task %s: %s", task_id, exc)
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="The automation service is currently unavailable. Please retry later.",
        )

    async def _sse_generator() -> AsyncGenerator[str, None]:
        """
        Async generator that reads from progress_queue and yields SSE events.

        SSE WIRE FORMAT
        ────────────────
        Each SSE "message" is: data: <payload>\n\n
        The double-newline terminates the message; the browser's EventSource
        parser splits on it to deliver individual events.
        """
        nonlocal future

        while True:
            # Poll the queue with a timeout so we also notice client disconnects.
            try:
                event = await asyncio.wait_for(progress_queue.get(), timeout=2.0)
            except asyncio.TimeoutError:
                # Check if the client has dropped the connection.
                if await request.is_disconnected():
                    logger.warning(
                        "[ENDPOINT] SSE client disconnected; cancelling job %s.", task_id
                    )
                    future.cancel()
                    return
                # Send a keep-alive comment so the TCP connection doesn't
                # time out on proxies / load-balancers.
                yield ": keep-alive\n\n"
                continue

            # None is the sentinel pushed by the Supervisor when the job ends.
            if event is None:
                break

            yield f"data: {json.dumps(event)}\n\n"

        # ------------------------------------------------------------------
        # All months done — await the final DataFrame from the Future.
        # ------------------------------------------------------------------
        try:
            result_df: pd.DataFrame = await asyncio.wait_for(
                future, timeout=60.0
            )
        except asyncio.TimeoutError:
            yield f"data: {json.dumps({'type': 'error', 'message': 'Timed out awaiting final result.'})}\n\n"
            return
        except Exception as exc:
            yield f"data: {json.dumps({'type': 'error', 'message': str(exc)})}\n\n"
            return

        if result_df is None or result_df.empty:
            yield f"data: {json.dumps({'type': 'complete', 'url': None, 'rows': 0, 'message': 'No data found for the requested period.'})}\n\n"
            return

        # Convert numeric month values to short names for the final Excel export.
        month_names = {
            1: "Jan",
            2: "Feb",
            3: "Mar",
            4: "Apr",
            5: "May",
            6: "Jun",
            7: "Jul",
            8: "Aug",
            9: "Sep",
            10: "Oct",
            11: "Nov",
            12: "Dec",
        }
        try:
            result_df = result_df.copy()
            if "month" in result_df.columns:
                result_df["month"] = result_df["month"].astype(int).map(month_names).fillna(result_df["month"])
            if "Month" in result_df.columns:
                result_df["Month"] = result_df["Month"].astype(int).map(month_names).fillna(result_df["Month"])
        except Exception:
            # If the month column cannot be converted cleanly, leave it as-is.
            pass

        # Serialize to .xlsx on a background thread with auto-sized columns.
        def _write_excel() -> bytes:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                result_df.to_excel(writer, index=False, sheet_name="Report")
                worksheet = writer.sheets["Report"]
                for i, col in enumerate(result_df.columns):
                    max_len = max(
                        result_df[col].astype(str).map(len).max() if not result_df[col].empty else 0, 
                        len(str(col))
                    )
                    from openpyxl.utils import get_column_letter
                    worksheet.column_dimensions[get_column_letter(i + 1)].width = max_len + 8
            return buf.getvalue()

        xlsx_bytes = await asyncio.to_thread(_write_excel)

        # Derive filename from the extracted data itself.
        party_name = _sanitize_filename_component(
            _pick_primary_party_name(result_df),
            fallback="party",
        )

        if "year" in result_df.columns:
            year_values = pd.to_numeric(result_df["year"], errors="coerce").dropna()
            year_tags = "_".join(str(int(year)) for year in sorted(year_values.astype(int).unique()))
        else:
            year_tags = "_".join(str(yc.year) for yc in body.years_config)

        filename = f"treasury_{party_name}_{year_tags}.xlsx" if year_tags else f"treasury_{party_name}.xlsx"

        saved_path = await asyncio.to_thread(save_dataframe_to_local_results, result_df, filename)
        saved_filename = os.path.basename(saved_path)

        # Open the saved file in Explorer so the Desktop view updates immediately.
        await asyncio.to_thread(_open_folder, saved_path)

        _store_download(session_id, xlsx_bytes, saved_filename)
        logger.info(
            "[ENDPOINT] Task %s saved Excel to %s and stored %d bytes under session_id=%s.",
            task_id, saved_path, len(xlsx_bytes), session_id,
        )

        yield f"data: {json.dumps({'type': 'complete', 'url': f'/api/v1/download/{session_id}', 'rows': len(result_df)})}\n\n"

    return StreamingResponse(
        content    = _sse_generator(),
        media_type = "text/event-stream",
        headers    = {
            "Cache-Control"               : "no-cache",
            "X-Accel-Buffering"           : "no",   # disable nginx buffering
            "Access-Control-Allow-Origin" : "*",
            "X-Task-Id"                   : task_id,
        },
    )


# ──────────────────────────────────────────────────────────────────────────────
# GET /download/{session_id}  →  fetch the compiled .xlsx after "complete"
# ──────────────────────────────────────────────────────────────────────────────

@router.get(
    "/download/{session_id}",
    summary="Fetch the compiled .xlsx after a completed extraction job",
    responses={
        200: {"content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}}},
        404: {"description": "Session not found or already expired"},
    },
)
async def download_result(session_id: str) -> Response:
    """
    One-shot file download endpoint.

    The client calls this after receiving the ``complete`` SSE event.
    The .xlsx bytes are served from _DOWNLOAD_STORE (in-memory).  The
    entry expires after _DOWNLOAD_TTL_SECONDS (30 min) and is pruned
    on the next write to _DOWNLOAD_STORE.
    """
    entry = _DOWNLOAD_STORE.get(session_id)
    if entry is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Download session not found or has expired. Please re-run the extraction.",
        )

    # Check TTL.
    if time.time() - entry["created_at"] > _DOWNLOAD_TTL_SECONDS:
        del _DOWNLOAD_STORE[session_id]
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Download session has expired. Please re-run the extraction.",
        )

    logger.info(
        "[ENDPOINT] GET /download/%s  → %d bytes  filename=%s",
        session_id, len(entry["bytes"]), entry["filename"],
    )

    return Response(
        content      = entry["bytes"],
        media_type   = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers      = {
            "Content-Disposition": f'attachment; filename="{entry["filename"]}"',
            "Content-Length":      str(len(entry["bytes"])),
        },
    )


# ──────────────────────────────────────────────────────────────────────────────
# GET /health
# ──────────────────────────────────────────────────────────────────────────────

@router.get(
    "/health",
    response_model=HealthResponse,
    summary="Supervisor health check",
)
async def health_check(
    supervisor: AutomationSupervisor = Depends(get_supervisor),
) -> HealthResponse:
    """
    Returns the current health of the AutomationSupervisor.

    The load balancer (Cloud Run / nginx) calls this endpoint every 30 seconds.
    If it returns 503, the instance is removed from the rotation.

    HTTP status mapping:
      • Circuit CLOSED → 200 OK
      • Circuit OPEN   → 503 Service Unavailable
    """
    h = supervisor.health

    if not h.is_healthy:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Automation Supervisor circuit breaker is OPEN.",
        )

    return HealthResponse(
        status              = "healthy",
        circuit_state       = h.circuit_state.name,
        queue_depth         = h.queue_depth,
        active_jobs         = h.active_jobs,
        consecutive_crashes = h.consecutive_crashes,
        total_jobs_done     = h.total_jobs_done,
        total_jobs_failed   = h.total_jobs_failed,
        uptime_seconds      = h.uptime_seconds,
    )


# ──────────────────────────────────────────────────────────────────────────────
# NOTE: _excel_stream_generator is no longer used for the primary /extract
# flow (which now uses SSE + /download).  It is retained here for any
# tooling or tests that may reference it directly.
# ──────────────────────────────────────────────────────────────────────────────

async def _excel_stream_generator(
    df: pd.DataFrame,
    chunk_size: int = 8_192,
) -> AsyncGenerator[bytes, None]:
    """Legacy helper — yields .xlsx bytes in chunks from a DataFrame."""
    def _write_excel_sync() -> io.BytesIO:
        buf = io.BytesIO()
        df.to_excel(buf, index=False, engine="openpyxl")
        buf.seek(0)
        return buf

    stream: io.BytesIO = await asyncio.to_thread(_write_excel_sync)
    while chunk := stream.read(chunk_size):
        yield chunk
        await asyncio.sleep(0)
